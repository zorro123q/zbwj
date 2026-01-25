from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook


class ExcelExporter:
    @staticmethod
    def export(result_json: Dict[str, Any], xlsx_path: Path) -> None:
        tables = result_json.get("tables") or []
        if not isinstance(tables, list) or len(tables) == 0:
            raise ValueError("no tables to export")

        wb = Workbook()

        # Remove default sheet if we will create our own named sheets
        default_ws = wb.active
        default_ws.title = "Sheet"

        created_any = False

        for idx, t in enumerate(tables):
            sheet_name = str(t.get("sheet_name") or f"Sheet{idx+1}")
            columns = t.get("columns") or []
            rows = t.get("rows") or []

            if not isinstance(columns, list) or len(columns) == 0:
                columns = ["col1"]
            if not isinstance(rows, list):
                rows = []

            if not created_any and default_ws.title == "Sheet":
                ws = default_ws
                ws.title = sheet_name[:31]
            else:
                ws = wb.create_sheet(title=sheet_name[:31])

            # header
            for c, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=c, value=str(col_name))

            # rows
            for r_idx, row in enumerate(rows, start=2):
                if not isinstance(row, list):
                    row = [row]
                for c, _ in enumerate(columns, start=1):
                    val = row[c - 1] if (c - 1) < len(row) else ""
                    ws.cell(row=r_idx, column=c, value=val)

            created_any = True

        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(xlsx_path))
