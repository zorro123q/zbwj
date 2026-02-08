# domain/review_index/requirements.py
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Any

from flask import current_app


@dataclass
class RequirementRow:
    category: str
    item: str
    value: str
    source: str


@dataclass
class AggregatedReq:
    """
    清洗后的聚合条目
    """
    category: str
    content_summary: str
    references: List[str] = field(default_factory=list)


def _repo_root() -> Path:
    return Path(current_app.root_path).parent


def load_requirements_xlsx(xlsx_path: str, sheet_name: str = "Result") -> List[RequirementRow]:
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required") from exc

    abs_path = Path(xlsx_path)
    if not abs_path.is_absolute():
        abs_path = (_repo_root() / abs_path).resolve()
    if not abs_path.exists():
        raise FileNotFoundError(f"xlsx not found: {abs_path}")

    wb = openpyxl.load_workbook(str(abs_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(x or "").strip().lower() for x in rows[0]]
    col = {name: header.index(name) for name in ["category", "item", "value", "source"] if name in header}

    # 稍微放宽限制，允许只有3列的情况（兼容性）
    if len(col) < 3:
        # 如果实在找不到 header，尝试按默认顺序
        col = {"category": 0, "item": 1, "value": 2, "source": 3}

    out: List[RequirementRow] = []
    for r in rows[1:]:
        # 确保行有内容
        if not r or not any(r):
            continue

        def _get(k, default_idx):
            idx = col.get(k, default_idx)
            if idx < len(r):
                return str(r[idx] or "").strip()
            return ""

        out.append(
            RequirementRow(
                category=_get("category", 0),
                item=_get("item", 1),
                value=_get("value", 2),
                source=_get("source", 3),
            )
        )
    return out


def _clean_source_to_page(source_text: str) -> str:
    """
    清洗来源：将 'line:120 ...' 转换为 'Page:3 ...'
    规则：页码 = 行号 // 50 + 1
    """
    if not source_text:
        return ""

    # 正则查找 line:数字 或 line 数字
    # group(1) 是数字
    def _repl(match):
        line_num = int(match.group(1))
        page_num = line_num // 50 + 1
        return f"Page:{page_num}"

    # 替换 "line:123" -> "Page:4"
    # re.IGNORECASE 忽略大小写
    cleaned = re.sub(r"line[:\s]*(\d+)", _repl, source_text, flags=re.IGNORECASE)
    return cleaned


def clean_and_aggregate_requirements(reqs: List[RequirementRow]) -> List[AggregatedReq]:
    """
    核心清洗逻辑：
    1. 忽略 item 里的 '条目1' 等后缀，直接按 category 聚合
    2. 合并 value 为大段文字
    3. 清洗 source 为 Page:X
    """
    from collections import defaultdict

    grouped = defaultdict(list)

    # 1. 分组
    for r in reqs:
        cat = r.category
        if not cat:
            cat = "其他"
        grouped[cat].append(r)

    results = []

    # 2. 聚合处理
    for cat, rows in grouped.items():
        # 合并内容：简单的换行拼接，或者用分号
        # 过滤掉空内容
        valid_values = [r.value for r in rows if r.value]
        summary = "\n".join(valid_values)

        # 处理来源：去重 + 格式化
        refs = set()
        for r in rows:
            clean_ref = _clean_source_to_page(r.source)
            if clean_ref:
                refs.add(clean_ref)

        # 排序来源，让 Page:1 排在 Page:10 前面 (简单的字符串排序可能不完美，但在展示上够用了)
        sorted_refs = sorted(list(refs))

        results.append(AggregatedReq(
            category=cat,
            content_summary=summary,
            references=sorted_refs
        ))

    return results


def requirements_to_kv(reqs: List[RequirementRow]) -> Dict[str, Dict[str, Any]]:
    """
    (保留原函数兼容性)
    """
    grouped: Dict[str, List[RequirementRow]] = {}
    for r in reqs:
        grouped.setdefault(r.category or "其他", []).append(r)

    return {
        cat: [{"item": x.item, "value": x.value, "source": x.source} for x in items]
        for cat, items in grouped.items()
    }